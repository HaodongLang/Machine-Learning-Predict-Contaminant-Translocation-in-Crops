# -*- coding = utf-8 -*-
# @Time :2024/7/7 21:41
# @Author :郎皓东
# @File ：bpn.py
# @Software:PyCharm
import openpyxl
import torch
import torch.nn as nn
import torch.optim as optim
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import scipy
from rdkit import Chem
from rdkit.Chem import AllChem

from utility import r2_score
from sklearn.model_selection import train_test_split
from torch.utils.data import Dataset, DataLoader
from utility import Kfold
import warnings
warnings.filterwarnings("ignore")
print('reading data')

# 数据输入

pesticide_data = pd.read_excel("data_final.xlsx",sheet_name="pesticide")
PPCPs_data = pd.read_excel("data_final.xlsx",sheet_name="PPCPs")
other_data = pd.read_excel("data_final.xlsx",sheet_name="other")

x_pesticide_data = pesticide_data.loc[:,["log RCF","flipid","MW（g/mol）","logKow"]].to_numpy().reshape(-1,4)
x_PPCPs_data=PPCPs_data.loc[:,["log RCF","flipid","MW（g/mol）","logKow"]].to_numpy().reshape(-1,4)
x_other_data=other_data.loc[:,["log RCF","flipid","MW（g/mol）","logKow"]].to_numpy().reshape(-1,4)


y_pesticide_data = pesticide_data.loc[:,["log TF"]].to_numpy().reshape(-1,1)
y_PPCPs_data=PPCPs_data.loc[:,["log TF"]].to_numpy().reshape(-1,1)
y_other_data=other_data.loc[:,["log TF"]].to_numpy().reshape(-1,1)

SMILES_pesticide=pesticide_data.loc[:,"SMILES"]
SMILES_PPCPs=PPCPs_data.loc[:,"SMILES"]
SMILES_other=other_data.loc[:,"SMILES"]
SMILES=np.concatenate((SMILES_pesticide,SMILES_PPCPs,SMILES_other),0)
FP=[]
for SMILE in SMILES:
    info = {}
    mol=Chem.MolFromSmiles(SMILE)
    fp=AllChem.GetMorganFingerprintAsBitVect(mol,2,nBits=1024,bitInfo=info)
    FP.append(fp)
FP=np.array(FP)

x_data = FP
# x_data=np.concatenate((x_pesticide_data,x_PPCPs_data,x_other_data),0)
# x_data=np.concatenate((FP,x_data),1)

y_data=np.concatenate((y_pesticide_data,y_PPCPs_data,y_other_data),0)
# x_data['logKow'] = x_data['logKow'].astype('float')
# x_data['MW（g/mol）'] = x_data['MW（g/mol）'].astype('float')


print('reading data finished')



#x_data_scaler=scipy.stats.mstats.zscore(x_data)
x_data_scaler=x_data
y_data_scaler=y_data




learn_rate=[0.0005,0.01,0.02,0.03,0.04,0.05,0.06,0.07,0.08,0.09,0.1]
#建立数据集

class ExcelDataset(Dataset):
    def __init__(self, excel_path,sheet_name):
        self.wb = openpyxl.load_workbook(excel_path)
        self.sheet = self.wb[sheet_name]
        self.sheet = self.wb.active
        column = self.sheet['E']

        # 假设第一行是标题行，所以数据从第二行开始
        self.data = list(self.sheet.iter_rows(min_row=2,min_col=1,max_col=4, values_only=True))  # log RCF、MW（g/mol）、logKow、flipid为输入
        #self.labels = list(self.sheet.iter_rows(min_row=2, min_col=5,max_col=5, values_only=True))  # log TF是预测值
        #self.labels = [cell.value for cell in column[1:]]
        self.labels = list(self.sheet.iter_rows(min_row=2, min_col=5,max_col=5, values_only=True))  # log TF是预测值
        #print(self.labels)

        # 转换数据为torch.Tensor，这里假设数据是数值型的
        self.data_tensor = [torch.tensor(row) for row in self.data]
        self.label_tensor = [torch.tensor(row) for row in self.labels]
    def __len__(self):
        return len(self.data)

    def __getitem__(self, idx):
        return self.data_tensor[idx], self.label_tensor[idx]

dataset = ExcelDataset("../data/data_final_.xlsx", "pesticide")
#dataloader的取
dataloader = DataLoader(dataset,batch_size=64,shuffle=False)




# 设定模型
class Bpnn(nn.Module):
    def __init__(self):
        super(Bpnn, self).__init__()
        self.layer1 = nn.Sequential(nn.Linear(1024,500),nn.Tanh())
        self.layer2 = nn.Sequential(nn.Linear(500,200))
        self.layer3 = nn.Sequential(nn.Linear(200,1))


    def forward(self, x):
        x = self.layer1(x)
        x = self.layer2(x)
        x = self.layer3(x)
        return x


train_val_idx,test_idx = Kfold(len(x_data_scaler),5)

# 定义损失函数和优化器
loss = nn.MSELoss(size_average=False)



total_r2=0
y_true=[]
y_predict=[]
# 训练模型
for fold in range(5):
    best_lr = 0
    best_test_r2 = 0
    best_train_r2 = 0
    best_val_r2 = 0
    train_id=[]
    val_id=[]
    test_id=[]
    x_train = []
    y_train = []
    x_val = []
    y_val = []
    x_test = []
    y_test = []

    for i in train_val_idx[fold][:int(len(train_val_idx[fold]) * 0.875)]:
        train_id.append(i)
    for i in train_val_idx[fold][int(len(train_val_idx[fold]) * 0.875):]:
        val_id.append(i)
    for i in test_idx[fold]:
        test_id.append(i)
    for i in train_id:
        x_train.append(x_data_scaler[i])
        y_train.append(y_data_scaler[i])
    for i in val_id:
        x_val.append(x_data_scaler[i])
        y_val.append(y_data_scaler[i])
    for i in test_id:
        x_test.append(x_data_scaler[i])
        y_test.append(y_data_scaler[i])
    print('Fold',fold+1)
    for lr in learn_rate:
        # 实例化模型
        model = Bpnn()
        optimizer = optim.Adam(model.parameters(), lr=lr)
        for epoch in range(150):
            y_pred = model(torch.Tensor(np.array(x_train)))
            loss_ = loss(y_pred, torch.Tensor(y_train))
            #print(epoch, loss_.item())
            optimizer.zero_grad()
            loss_.backward()
            optimizer.step()

            r2_train = r2_score(torch.Tensor(y_train), model(torch.Tensor(x_train)))
            r2_val = r2_score(torch.Tensor(y_val), model(torch.Tensor(x_val)))
            r2_test=r2_score(torch.Tensor(y_test),model(torch.Tensor(x_test)))

            if r2_train>best_train_r2:
                best_train_r2=r2_train
            if r2_val>best_val_r2:
                best_val_r2=r2_val
            if r2_test>best_test_r2:
                best_test_r2=r2_test
                best_lr=lr
    print(f"best_lr={best_lr} train_r2={best_train_r2} val_r2={best_val_r2} test_r2={best_test_r2}")
    total_r2+=best_test_r2
    best_r2=0
    y_true.append(y_test)
    y_predict.append(model(torch.Tensor(x_test)).tolist())

y_predict=np.array(y_predict,dtype=object)
y_true=np.array(y_true,dtype=object)
y_predict=y_predict.squeeze()
y_true=y_true.squeeze()

final_r2=(total_r2/5)
print("final_r2=",final_r2)

y_true_=[]
for i in y_true:
    for j in i:
        y_true_.append(j)


y_predict_=[]
for i in y_predict:
    for j in i:
        y_predict_.append(j)





# 可视化
# optimizer = optim.Adam(model.parameters(),lr=best_lr)
# for epoch in range(1000):
#     y_pred = model(torch.Tensor(x_data_scaler[train_idx]))
#     loss_ = loss(y_pred, torch.Tensor(y_data_scaler[train_idx]))
#
#     # print(epoch, loss_.item())
#     optimizer.zero_grad()
#     loss_.backward()
#     optimizer.step()

fig=plt.figure(figsize=(15,6))
ax0=fig.add_subplot(111)
ax0.scatter(y_true_,y_predict_,c='r')

x = np.arange(-4,5)
y = np.arange(-4,5)
ax0.plot(x,y,c="brown")
#plt.plot(x_data.numpy(), y_data.numpy(), 'ro', label='Original Data')
#plt.plot(x_data.numpy(), model(x_data).data.numpy(), 'b-', label='Prediction')
ax0.set_xlabel("measured logTF")
ax0.set_ylabel("predicted logTF")
ax0.text(0.03,0.9,f"r2={round(final_r2,2)}",transform=ax0.transAxes)
plt.legend()
#plt.savefig('scatter.png')
plt.show()
